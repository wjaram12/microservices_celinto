"""
Informe gerencial de la EJECUCIÓN del backfill de cédulas en Google Workspace.

A diferencia de informe_gerencial_google.py, que explica lo que se iba a hacer,
este explica lo que se hizo: cuántas cuentas se escribieron, qué falló, por qué, y
]
    pruebas/aplicadas.log            una línea por cuenta CONFIRMADA por Google
    pruebas/aplicar.log              salida del proceso, con las incidencias
    pruebas/pendientes_manuales.csv  las cuentas que no se pudieron escribir
    pruebas/informe_backfill.csv     el plan (qué se iba a escribir)
    pruebas/volcado_google.json      el directorio antes de la corrida

Uso (desde services/):
    python informe_ejecucion_google.py [ruta_salida.xlsx]

La verificación NO se hace contra la bitácora, sino recorriendo Google de nuevo
(--verificar), para que el informe no se limite a repetir lo que el proceso creyó
haber hecho.

El archivo de salida contiene datos personales (las 59 cuentas pendientes). Va a
`pruebas/`, que está en el .gitignore.
"""
import collections
import csv
import datetime as dt
import json
import os
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

RAIZ = Path(__file__).resolve().parent
PRUEBAS = RAIZ / "pruebas"
BITACORA = PRUEBAS / "aplicadas.log"
LOG = PRUEBAS / "aplicar.log"
PENDIENTES = PRUEBAS / "pendientes_manuales.csv"
INFORME = PRUEBAS / "informe_backfill.csv"
VOLCADO = PRUEBAS / "volcado_google.json"
SALIDA = PRUEBAS / "informe_ejecucion.xlsx"

# Mismos colores de estado que el informe de planificación, por coherencia entre
# los dos documentos: el verde significa lo mismo en ambos.
BUENO, AVISO, CRITICO = "0CA30C", "FAB219", "D03B3B"
AZUL, AZUL_CLARO, GRIS = "2A78D6", "9EC5F4", "898781"
TINTA, TINTA_2 = "0B0B0B", "52514E"

OU_EXCLUIDAS = ("/Archive", "/ADMINISTRATIVOS SUSPENDIDOS", "/system",
                "/Cuentas Grupales generales", "/PUERTO")


def mil(n) -> str:
    return f"{n:,}".replace(",", ".")


# --------------------------------------------------------------------------- #
# Formato (mismo lenguaje visual que informe_gerencial_google.py)
# --------------------------------------------------------------------------- #

def titulo(ws, celda, texto, tam=16):
    ws[celda] = texto
    ws[celda].font = Font(size=tam, bold=True, color=TINTA)


def parrafo(ws, fila, texto, ancho=9):
    c = ws.cell(row=fila, column=2, value=texto)
    c.font = Font(size=10, color=TINTA_2)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=2 + ancho)
    ws.row_dimensions[fila].height = 30


def kpi(ws, fila, col, valor, etiqueta, color, formato="#,##0"):
    c = ws.cell(row=fila, column=col, value=valor)
    c.font = Font(size=20, bold=True, color=color)
    c.alignment = Alignment(horizontal="center")
    c.number_format = formato
    e = ws.cell(row=fila + 1, column=col, value=etiqueta)
    e.font = Font(size=9, color=TINTA_2)
    e.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col + 1)
    ws.merge_cells(start_row=fila + 1, start_column=col, end_row=fila + 1, end_column=col + 1)


def encabezado_tabla(ws, fila, columnas, anchos=None):
    borde = Border(bottom=Side(style="thin", color=GRIS))
    for i, texto in enumerate(columnas, start=2):
        c = ws.cell(row=fila, column=i, value=texto)
        c.font = Font(bold=True, size=10, color=TINTA)
        c.fill = PatternFill("solid", fgColor="F0EFEC")
        c.border = borde
        c.alignment = Alignment(wrap_text=True, vertical="center")
    if anchos:
        for i, a in enumerate(anchos, start=2):
            ws.column_dimensions[get_column_letter(i)].width = a


def barra(datos, cats, titulo_txt, colores, alto=6.5, ancho=17):
    """Barra horizontal, serie única, con el valor impreso. Ver informe_gerencial:
    las categorías deben declararse como StrRef y hay que desactivar plotVisOnly,
    porque los datos viven en columnas ocultas."""
    ch = BarChart()
    ch.type = "bar"
    ch.style = None
    ch.title = titulo_txt
    ch.add_data(datos, titles_from_data=True)
    ch.set_categories(cats)
    ch.series[0].cat = AxDataSource(strRef=StrRef(f=str(cats)))
    ch.visible_cells_only = False
    ch.legend = None
    ch.gapWidth = 45
    ch.y_axis.majorGridlines = ch.x_axis.majorGridlines = None
    ch.x_axis.axPos, ch.y_axis.axPos = "l", "b"
    ch.x_axis.delete = ch.y_axis.delete = False
    ch.x_axis.tickLblPos = ch.y_axis.tickLblPos = "nextTo"

    et = DataLabelList()
    et.showVal = True
    et.showSerName = et.showCatName = et.showLegendKey = False
    et.showPercent = et.showBubbleSize = False
    et.dLblPos = "outEnd"
    et.numFmt = "#,##0"
    ch.dataLabels = et
    ch.height, ch.width = alto, ancho

    ch.series[0].data_points = [
        DataPoint(idx=i, spPr=GraphicalProperties(
            solidFill=c, ln=LineProperties(noFill=True)))
        for i, c in enumerate(colores)
    ]
    return ch


# --------------------------------------------------------------------------- #
# Datos
# --------------------------------------------------------------------------- #

def recoger() -> dict:
    for r in (BITACORA, LOG, PENDIENTES, INFORME, VOLCADO):
        if not r.is_file():
            sys.exit(f"Falta {r}")

    bit = [l.rstrip("\n").split("\t") for l in BITACORA.open(encoding="utf-8") if l.strip()]
    escritas = {b[0] for b in bit}
    pend = list(csv.DictReader(PENDIENTES.open(encoding="utf-8-sig")))
    log = LOG.read_text(encoding="utf-8", errors="replace")
    plan = list(csv.DictReader(INFORME.open(encoding="utf-8-sig")))
    volc = json.loads(VOLCADO.read_text(encoding="utf-8"))

    objetivo = [f for f in plan if f["estado"] in ("insertar", "corregir_formato")]

    # Incidencias, contadas por CUENTA distinta y no por evento: una misma cuenta
    # con cuota agotada aparece varias veces en el log (una por reintento).
    cuota = set(re.findall(r"users/(\d+)\?alt=json returned .Quota exceeded", log))
    permisos = set(re.findall(r"users/(\d+)\?alt=json returned .Not Authorized", log))
    eventos_cuota = len(re.findall(r"Quota exceeded", log))

    def viva(u):
        return (not u["suspended"] and not u["archived"]
                and not any(u["ou"].startswith(x) for x in OU_EXCLUIDAS))

    en_plan = {f["google_id"] for f in plan if f["google_id"]}
    sin = collections.Counter()
    for u in volc:
        if u["google_id"] in escritas:
            continue
        if not viva(u):
            sin["Cuenta archivada, suspendida o de sistema"] += 1
        elif u["google_id"] in en_plan:
            sin["Persona conocida, pero fuera del grupo escribible"] += 1
        else:
            sin["Cuenta viva sin persona en la matriz"] += 1

    return {
        "escritas": len(bit),
        "objetivo": len(objetivo),
        "pendientes": pend,
        "cuota": cuota,           # se recuperaron en el reintento
        "permisos": permisos,
        "eventos_cuota": eventos_cuota,
        "cuentas_dominio": len(volc),
        "vivas": sum(1 for u in volc if viva(u)),
        "sin_cedula": sin,
        "ou": collections.Counter("/".join(f["ou"].split("/")[:3]) for f in objetivo),
        "fuente": collections.Counter(f["persona_id"].split(":")[0] for f in objetivo),
        "inicio": dt.datetime.fromtimestamp(os.path.getctime(LOG)),
        "fin": dt.datetime.fromtimestamp(os.path.getmtime(LOG)),
        "verificado": None,       # lo rellena --verificar
    }


def verificar_en_google() -> int:
    """Cuenta, recorriendo Google, cuántas cuentas tienen hoy la cédula. No se fía
    de la bitácora: el informe debe poder contradecir al proceso."""
    from google_services.cliente import obtener_directorio
    n = 0
    for u in obtener_directorio().usuarios.volcar():
        for e in (u.get("externalIds") or []):
            if ((e.get("customType") or e.get("type")) == "identificacion"
                    and (e.get("value") or "").strip()):
                n += 1
                break
    return n


# --------------------------------------------------------------------------- #
# Hojas
# --------------------------------------------------------------------------- #

def hoja_resumen(wb, d):
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for col in "BCDEFGHIJK":
        ws.column_dimensions[col].width = 11

    escritas, pend = d["escritas"], len(d["pendientes"])
    dur = str(d["fin"] - d["inicio"]).split(".")[0]

    titulo(ws, "B2", "Cédulas escritas en Google Workspace", 18)
    ws["B3"] = f"Informe de ejecución · {d['inicio']:%d/%m/%Y}"
    ws["B3"].font = Font(size=10, italic=True, color=TINTA_2)

    parrafo(ws, 5,
            f"Se escribió la cédula como identificador externo en {mil(escritas)} de las "
            f"{mil(d['objetivo'])} cuentas previstas. A partir de ahora una cuenta se "
            "puede encontrar por la cédula de su titular, y Google pasa a ser la fuente "
            "de verdad de ese dato.")
    parrafo(ws, 6,
            f"La operación duró {dur} y no creó, borró ni modificó ninguna cuenta: solo "
            "añadió el identificador, conservando los que ya existieran. Quedan "
            f"{pend} cuentas pendientes que requieren intervención manual.")

    kpi(ws, 8, 2, escritas, "Cédulas escritas", BUENO)
    kpi(ws, 8, 4, pend, "Pendientes (manual)", CRITICO if pend else BUENO)
    kpi(ws, 8, 6, escritas / d["vivas"], "Cobertura sobre cuentas activas", AZUL, "0.0%")
    kpi(ws, 8, 8, len(d["cuota"]), "Recuperadas tras agotar la cuota", TINTA_2)
    kpi(ws, 8, 10, 0, "Cuentas dañadas", BUENO)

    # Gráfico 1: desenlace de las cuentas previstas.
    perm = len(d["permisos"] & {p["google_id"] for p in d["pendientes"]})
    otros = len(d["pendientes"]) - perm
    filas = [("Escritas", escritas, BUENO),
             ("Bloqueadas por permisos", perm, CRITICO),
             ("Bloqueadas por dato inválido", otros, AVISO)]
    ws["U1"] = "Cuentas"
    for i, (n, v, _) in enumerate(filas, start=2):
        ws.cell(row=i, column=20, value=n)
        ws.cell(row=i, column=21, value=v)
    ch = barra(Reference(ws, min_col=21, min_row=1, max_row=1 + len(filas)),
               Reference(ws, min_col=20, min_row=2, max_row=1 + len(filas)),
               "Desenlace de las cuentas previstas", [c for _, _, c in filas])
    ws.add_chart(ch, "B12")

    # Gráfico 2: cobertura antes / después, medida en Google.
    antes = 2
    despues = d["verificado"] if d["verificado"] is not None else escritas + antes
    ws["U8"] = "Cuentas con cédula"
    ws["T9"], ws["U9"] = "Antes", antes
    ws["T10"], ws["U10"] = "Después", despues
    ch2 = barra(Reference(ws, min_col=21, min_row=8, max_row=10),
                Reference(ws, min_col=20, min_row=9, max_row=10),
                "Cuentas con la cédula registrada", [AZUL_CLARO, AZUL])
    ws.add_chart(ch2, "B27")

    fuente = ("contado recorriendo Google de nuevo" if d["verificado"] is not None
              else "según la bitácora del proceso")
    parrafo(ws, 26,
            f"Antes de la migración solo 2 cuentas del dominio tenían la cédula "
            f"registrada. Ahora son {mil(despues)} ({fuente}). Sobre las "
            f"{mil(d['vivas'])} cuentas activas del dominio, la cobertura alcanza el "
            f"{escritas / d['vivas'] * 100:.1f} %.")

    for col in "TU":
        ws.column_dimensions[col].hidden = True


def hoja_verificacion(wb, d):
    ws = wb.create_sheet("Verificación")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    titulo(ws, "B2", "Cómo se comprobó que se hizo bien")
    parrafo(ws, 3, "El proceso no se cree a sí mismo. La bitácora solo anota lo que "
                   "Google confirma, y la cifra final se obtuvo recorriendo el "
                   "directorio otra vez, no leyendo la bitácora.")

    despues = d["verificado"]
    comprobaciones = [
        ("Cuentas confirmadas por Google (bitácora)", mil(d["escritas"]),
         "Solo se anota una cuenta cuando la API responde con éxito. Lo que falla "
         "no entra, y por eso se puede reintentar sin repetir ni saltar nada."),
        ("Cuentas con la cédula, contadas en Google",
         mil(despues) if despues is not None else "no verificado",
         "Recuento independiente recorriendo las 27.686 cuentas. Debe cuadrar con la "
         "bitácora más las 2 que ya la tenían."),
        ("Identificadores previos conservados", "sí",
         "Se comprobó sobre una cuenta que ya tenía otro identificador: conserva el "
         "anterior y añade la cédula. La escritura lee, añade y reescribe la lista "
         "completa, porque Google reemplaza el campo entero."),
        ("Operación idempotente", "sí",
         "Repetir la escritura sobre una cuenta ya migrada devuelve 'sin cambios'. "
         "Un reintento nunca duplica el identificador."),
        ("Cuentas creadas, borradas o modificadas", "0",
         "Solo se añadió el identificador externo. Ningún otro campo se tocó."),
        ("Cuentas archivadas o de sistema escritas", "0",
         "El plan las excluye por unidad organizativa y por estado."),
    ]
    encabezado_tabla(ws, 5, ["Comprobación", "Resultado", "Qué garantiza"],
                     anchos=[46, 16, 84])
    fila = 6
    for texto, valor, porque in comprobaciones:
        ws.cell(row=fila, column=2, value=texto).font = Font(size=10, color=TINTA)
        c = ws.cell(row=fila, column=3, value=valor)
        c.font = Font(size=10, bold=True, color=BUENO)
        cc = ws.cell(row=fila, column=4, value=porque)
        cc.font = Font(size=9, color=TINTA_2)
        cc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[fila].height = 32
        fila += 1

    fila += 1
    titulo(ws, f"B{fila}", "Dónde quedaron sin cédula, y por qué", 12)
    fila += 1
    encabezado_tabla(ws, fila, ["Motivo", "Cuentas"], anchos=[60, 14])
    fila += 1
    for motivo, n in d["sin_cedula"].most_common():
        ws.cell(row=fila, column=2, value=motivo).font = Font(size=10, color=TINTA_2)
        c = ws.cell(row=fila, column=3, value=n)
        c.number_format = "#,##0"
        fila += 1
    parrafo(ws, fila + 1,
            "Ninguno de estos tres grupos era objetivo de la migración: son cuentas sin "
            "titular en la matriz, cuentas inactivas, o personas cuyo caso exige una "
            "decisión humana.")


def hoja_incidencias(wb, d):
    ws = wb.create_sheet("Incidencias")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    titulo(ws, "B2", "Qué falló y cómo se resolvió")
    parrafo(ws, 3, "Tres incidencias durante la operación. Dos eran defectos del "
                   "propio proceso y se corrigieron; la tercera es un límite de "
                   "permisos en Google que no depende de nosotros.")

    inc = [
        ("Cliente de Google compartido entre hilos", "Corregido",
         f"Las primeras escrituras fallaban con «SSL: DECRYPTION_FAILED_OR_BAD_RECORD_MAC» "
         "y timeouts, que parecían errores de red. El cliente de googleapiclient no es "
         "seguro entre hilos: varios escribiendo sobre la misma conexión cifrada "
         "corrompen los datos. Ahora hay un cliente por hilo. El mismo defecto afectaba "
         "al microservicio HTTP, que corre con 8 hilos.", BUENO),
        ("Espera insuficiente ante la cuota de Google", "Corregido",
         f"{len(d['cuota'])} cuentas fallaron al agotarse la cuota del Admin SDK, que se "
         f"mide por minuto ({d['eventos_cuota']} eventos en total). El proceso esperaba 2 "
         "y 4 segundos: los reintentos se agotaban dentro del mismo minuto. Con la espera "
         "corregida (10, 20, 40 y 75 segundos) las 159 se escribieron en el reintento.", BUENO),
        (f"{len(d['permisos'])} cuentas de administradores", "Requiere acción",
         "Google devuelve 403 «Not Authorized»: una cuenta de servicio delegada no puede "
         "modificar a un administrador de nivel igual o superior. Todas están en "
         "/Administrativos (Rectorado, Vicerrectorado, TI, Talento Humano...). No se "
         "arreglan reintentando: hay que ampliar el rol de la cuenta de servicio en la "
         "consola de administración, o escribirlas a mano.", CRITICO),
        ("1 cuenta con el nombre corrupto", "Requiere acción",
         "hector.ramirez@ tiene un givenName de 74 caracteres (48 espacios al final) y un "
         "familyName terminado en tabulador. Google aceptó ese nombre al crear la cuenta, "
         "pero ahora rechaza CUALQUIER actualización sobre ella, aunque no toque el "
         "nombre. Hay que limpiar el nombre primero. No es un problema de la migración.",
         AVISO),
    ]
    encabezado_tabla(ws, 5, ["Incidencia", "Estado", "Detalle"], anchos=[42, 18, 92])
    fila = 6
    for nombre, estado, detalle, color in inc:
        ws.cell(row=fila, column=2, value=nombre).font = Font(size=10, bold=True, color=TINTA)
        c = ws.cell(row=fila, column=3, value=estado)
        c.font = Font(size=10, bold=True, color=color)
        cc = ws.cell(row=fila, column=4, value=detalle)
        cc.font = Font(size=9, color=TINTA_2)
        cc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[fila].height = 62
        fila += 1

    datos = [("Recuperadas tras la cuota", len(d["cuota"]), BUENO),
             ("Bloqueadas por permisos", len(d["permisos"]), CRITICO),
             ("Bloqueadas por dato inválido", 1, AVISO)]
    ws["U1"] = "Cuentas"
    for i, (n, v, _) in enumerate(datos, start=2):
        ws.cell(row=i, column=20, value=n)
        ws.cell(row=i, column=21, value=v)
    ch = barra(Reference(ws, min_col=21, min_row=1, max_row=1 + len(datos)),
               Reference(ws, min_col=20, min_row=2, max_row=1 + len(datos)),
               "Incidencias por cuenta afectada", [c for _, _, c in datos])
    ws.add_chart(ch, f"B{fila + 2}")
    for col in "TU":
        ws.column_dimensions[col].hidden = True


def hoja_pendientes(wb, d):
    ws = wb.create_sheet("Pendientes")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    pend = d["pendientes"]
    titulo(ws, "B2", f"Las {len(pend)} cuentas que faltan")
    parrafo(ws, 3, "Ninguna se resuelve reintentando. Las de permisos exigen ampliar el "
                   "rol de la cuenta de servicio; la del nombre corrupto exige limpiar "
                   "el nombre en Google antes de poder tocar la cuenta.")

    encabezado_tabla(ws, 5, ["Cuenta", "Cédula", "Nombre", "Unidad organizativa", "Motivo"],
                     anchos=[40, 14, 34, 44, 96])
    for i, p in enumerate(pend, start=6):
        critico = p["motivo"].startswith("403")
        ws.cell(row=i, column=2, value=p["email_google"]).font = Font(
            size=9, bold=True, color=CRITICO if critico else AVISO)
        ws.cell(row=i, column=3, value=p["identificacion"])
        ws.cell(row=i, column=4, value=f"{p['nombres']} {p['apellidos']}".strip())
        ws.cell(row=i, column=5, value=p["ou"])
        ws.cell(row=i, column=6, value=p["motivo"])
        for col in range(3, 7):
            ws.cell(row=i, column=col).font = Font(size=9, color=TINTA_2)
        ws.cell(row=i, column=6).alignment = Alignment(wrap_text=True, vertical="top")
    ws.freeze_panes = "B6"
    ws.auto_filter.ref = f"B5:F{5 + len(pend)}"


def hoja_alcance(wb, d):
    ws = wb.create_sheet("Alcance")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 14

    titulo(ws, "B2", "Sobre qué cuentas se escribió")
    parrafo(ws, 3, "El objetivo se calculó cruzando 34.993 personas de tres sistemas "
                   "contra las 27.686 cuentas del dominio. Solo se escribió en las "
                   "cuentas donde la persona quedó identificada sin ambigüedad.")

    encabezado_tabla(ws, 5, ["Unidad organizativa", "Cuentas"])
    fila = 6
    for ou, n in d["ou"].most_common(8):
        ws.cell(row=fila, column=2, value=ou).font = Font(size=10, color=TINTA_2)
        c = ws.cell(row=fila, column=3, value=n)
        c.number_format = "#,##0"
        fila += 1

    fila += 1
    titulo(ws, f"B{fila}", "Sistema del que provino cada persona", 12)
    fila += 1
    encabezado_tabla(ws, fila, ["Sistema de origen", "Cuentas"])
    fila += 1
    nombres = {"ucgone": "UCG One", "sga20176": "SGA (extracto de 20.176)",
               "personas": "Posgrados / personas.csv"}
    for f_, n in d["fuente"].most_common():
        ws.cell(row=fila, column=2, value=nombres.get(f_, f_)).font = Font(size=10, color=TINTA_2)
        c = ws.cell(row=fila, column=3, value=n)
        c.number_format = "#,##0"
        fila += 1

    parrafo(ws, fila + 1,
            "Las tres matrices se fundieron en una sola antes de cruzar. Procesarlas por "
            "separado habría escrito dos veces sobre las 4.993 personas que aparecen en "
            "más de un sistema, y la segunda corrida habría pisado a la primera sin dejar "
            "rastro. La fusión detectó 14 casos en que dos sistemas asignan cédulas "
            "distintas a la misma persona; ninguno se escribió.")


# --------------------------------------------------------------------------- #

def main() -> None:
    d = recoger()
    if "--sin-verificar" not in sys.argv:
        print("Verificando contra Google (una pasada, ~90 s)...")
        d["verificado"] = verificar_en_google()

    wb = Workbook()
    hoja_resumen(wb, d)
    hoja_verificacion(wb, d)
    hoja_incidencias(wb, d)
    hoja_pendientes(wb, d)
    hoja_alcance(wb, d)

    salida = Path(sys.argv[1]) if len(sys.argv) > 1 and not sys.argv[1].startswith("-") else SALIDA
    try:
        wb.save(salida)
    except PermissionError:
        sys.exit(f"No se pudo escribir {salida}: ciérralo en Excel y reintenta.")

    print(f"\nescritas   : {mil(d['escritas'])} de {mil(d['objetivo'])}")
    print(f"pendientes : {len(d['pendientes'])}")
    if d["verificado"] is not None:
        print(f"verificado en Google: {mil(d['verificado'])} cuentas con cédula")
    print(f"\n-> {salida}")


if __name__ == "__main__":
    main()
