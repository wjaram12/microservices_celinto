"""
Genera un informe gerencial en Excel a partir del informe técnico del backfill.

Lee `pruebas/informe_backfill.csv` (producido por migrar_cedulas_google.py) y, si
está, `pruebas/volcado_google.json` para el diagnóstico del directorio. Escribe
`pruebas/informe_gerencial.xlsx` con cinco hojas: Resumen, Estados, Calidad del
dato, Revisión y Metodología.

Uso (desde services/):
    python informe_gerencial_google.py


Sobre los gráficos. Con once estados distintos, una tarta sería ilegible: por
encima de ~7 clases la recomendación es tabla, no más colores. Así que la vista
principal agrupa los estados en CUATRO acciones y el detalle vive en una tabla.
Cada barra lleva su valor impreso, de modo que el color nunca es el único canal
que transporta la información.

El archivo de salida contiene datos personales (nombres y cédulas de las filas que
requieren revisión). Va a `pruebas/`, que está en el .gitignore.
"""
import csv
import json
import sys
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.data_source import AxDataSource, StrRef
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.line import LineProperties
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    sys.stdout.reconfigure(encoding="utf-8")
except (AttributeError, ValueError):
    pass

RAIZ = Path(__file__).resolve().parent
PRUEBAS = RAIZ / "pruebas"
INFORME = PRUEBAS / "informe_backfill.csv"
VOLCADO = PRUEBAS / "volcado_google.json"
SALIDA = PRUEBAS / "informe_gerencial.xlsx"

# --- Paleta -----------------------------------------------------------------
# Colores de ESTADO (reservados, nunca reutilizados como "una serie más"). Cada
# barra lleva impreso su valor y existe la tabla completa en la hoja "Estados":
# el color nunca es el único canal que transporta la información. Esa es la
# condición que permite usar el ámbar, que a propósito queda por debajo de 3:1
# de contraste sobre fondo claro.
BUENO = "0CA30C"     # listo, sin intervención
AVISO = "FAB219"     # requiere decisión humana
CRITICO = "D03B3B"   # riesgo de escribir un dato equivocado
AZUL = "2A78D6"      # magnitud neutra (hue secuencial por defecto)
AZUL_CLARO = "9EC5F4"
GRIS = "898781"      # sin acción / contexto (gris de énfasis, >=3:1 sobre el fondo)

TINTA = "0B0B0B"
TINTA_2 = "52514E"
SUPERFICIE = "FCFCFB"
LINEA = "E1E0D9"

# --- Catálogo de estados ----------------------------------------------------
# (grupo de acción, color, qué significa, qué hay que hacer)
ESTADOS = {
    "insertar": ("Listas para escribir", BUENO,
                 "Se identificó su cuenta principal y no tiene la cédula registrada.",
                 "Ninguna: la escribe el proceso automático."),
    "corregir_formato": ("Listas para escribir", BUENO,
                         "Tiene su cédula pero mal escrita (le falta el cero inicial).",
                         "Ninguna: el proceso la normaliza."),
    "vinculada": ("Ya vinculadas", GRIS,
                  "Su cédula ya está correctamente registrada en Google.",
                  "Ninguna."),
    "disponible": ("Sin cuenta en Google", AZUL,
                   "La persona no tiene ninguna cuenta en el dominio.",
                   "Decidir si corresponde crearle una cuenta."),
    "correo_ocupado": ("Requieren revisión", AVISO,
                       "El correo que figura en la tabla pertenece a OTRA persona.",
                       "Corregir el correo en la tabla de personas."),
    "ambigua": ("Requieren revisión", AVISO,
                "Varias cuentas activas con el mismo nombre y la misma jerarquía.",
                "Elegir a mano cuál es la cuenta de la persona."),
    "revisar_multicuenta": ("Requieren revisión", AVISO,
                            "La persona tiene dos cuentas (p. ej. docente y estudiante) "
                            "y la principal no es la que la identificó.",
                            "Confirmar en cuál debe ir la cédula."),
    "revisar_nombre_difuso": ("Requieren revisión", AVISO,
                              "Solo se encontró un nombre parecido, no idéntico.",
                              "Confirmar que es la misma persona."),
    "solo_cuentas_inactivas": ("Requieren revisión", AVISO,
                               "Sus únicas cuentas están archivadas o suspendidas.",
                               "Decidir si se reactiva o se crea una nueva."),
    "conflicto_cedula": ("Requieren revisión", CRITICO,
                         "Google tiene registrada una cédula DISTINTA de la de la tabla.",
                         "Averiguar cuál es la correcta. No se sobrescribe nada."),
    "conflicto_duplicado": ("Requieren revisión", CRITICO,
                            "Dos personas de la tabla apuntan a la misma cuenta con "
                            "cédulas distintas.",
                            "Depurar la tabla de personas."),
    "duplicado_persona": ("Requieren revisión", AVISO,
                          "La misma persona aparece dos veces en la tabla, con la "
                          "misma cédula.",
                          "Eliminar la fila repetida."),
    "cedula_invalida": ("Requieren revisión", CRITICO,
                        "La cédula es un valor de relleno (p. ej. 0000000000).",
                        "Cargar la cédula real antes de migrar."),
}

ORDEN_GRUPOS = ["Listas para escribir", "Sin cuenta en Google", "Requieren revisión",
                "Ya vinculadas"]
COLOR_GRUPO = {"Listas para escribir": BUENO, "Sin cuenta en Google": AZUL,
               "Requieren revisión": AVISO, "Ya vinculadas": GRIS}
REVISION = [e for e, v in ESTADOS.items() if v[0] == "Requieren revisión"]


# --- Utilidades de formato --------------------------------------------------

def mil(n) -> str:
    """27686 -> '27.686'. Solo sobre el número: aplicar el reemplazo a la frase
    entera convertiría también las comas de la redacción."""
    return f"{n:,}".replace(",", ".")


def titulo(ws, celda, texto, tam=16):
    ws[celda] = texto
    ws[celda].font = Font(size=tam, bold=True, color=TINTA)


def parrafo(ws, fila, texto, ancho=9):
    ws.cell(row=fila, column=2, value=texto)
    ws.cell(row=fila, column=2).font = Font(size=10, color=TINTA_2)
    ws.cell(row=fila, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=2 + ancho)
    ws.row_dimensions[fila].height = 30


def kpi(ws, fila, col, valor, etiqueta, color):
    c = ws.cell(row=fila, column=col, value=valor)
    c.font = Font(size=20, bold=True, color=color)
    c.alignment = Alignment(horizontal="center")
    c.number_format = "#,##0"
    e = ws.cell(row=fila + 1, column=col, value=etiqueta)
    e.font = Font(size=9, color=TINTA_2)
    e.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.merge_cells(start_row=fila, start_column=col, end_row=fila, end_column=col + 1)
    ws.merge_cells(start_row=fila + 1, start_column=col, end_row=fila + 1, end_column=col + 1)


def encabezado_tabla(ws, fila, columnas, anchos=None):
    borde = Border(bottom=Side(style="thin", color=GRIS))
    for i, texto in enumerate(columnas, start=2):
        c = ws.cell(row=fila, column=i, value=texto)
        c.font = Font(bold=True, size=10, color=TINTA)
        c.fill = PatternFill("solid", fgColor="F0EFEC")
        c.border = borde
        c.alignment = Alignment(wrap_text=True, vertical="center")
    if anchos:
        for i, a in enumerate(anchos, start=2):
            ws.column_dimensions[get_column_letter(i)].width = a


def barra(datos_ref, cats_ref, titulo_txt, colores, alto=7.5, ancho=16):
    """Barra horizontal, serie única, con el valor impreso en cada barra.

    Serie única -> sin leyenda (el título ya la nombra). Sin rejilla: la etiqueta
    numérica hace innecesario el eje de valores como referencia visual.
    """
    ch = BarChart()
    ch.type = "bar"
    ch.style = None
    ch.title = titulo_txt
    ch.add_data(datos_ref, titles_from_data=True)
    ch.set_categories(cats_ref)

    # Las categorías son TEXTO. `set_categories` las declara como referencia
    # numérica, y Excel entonces rotula las barras 1, 2, 3... en vez de con su
    # nombre. Hay que declararlas explícitamente como referencia de cadena.
    ch.series[0].cat = AxDataSource(strRef=StrRef(f=str(cats_ref)))

    # Los datos viven en columnas ocultas. Por defecto Excel NO grafica celdas
    # ocultas (plotVisOnly), y el gráfico saldría vacío.
    ch.visible_cells_only = False

    ch.legend = None
    ch.gapWidth = 45
    ch.y_axis.majorGridlines = None
    ch.x_axis.majorGridlines = None

    # En una barra horizontal el eje de categorías va a la izquierda y el de
    # valores abajo; openpyxl deja los dos a la izquierda. Sin `delete=False` y
    # `tickLblPos`, Excel puede no dibujar los rótulos.
    ch.x_axis.axPos, ch.y_axis.axPos = "l", "b"
    ch.x_axis.delete, ch.y_axis.delete = False, False
    ch.x_axis.tickLblPos = ch.y_axis.tickLblPos = "nextTo"

    # Solo el VALOR. Si los demás canales se dejan sin declarar, Excel los activa y
    # cada barra queda rotulada "Personas; Listas para escribir; 14276" con su marca
    # de leyenda al lado.
    etiquetas = DataLabelList()
    etiquetas.showVal = True
    etiquetas.showSerName = False
    etiquetas.showCatName = False
    etiquetas.showLegendKey = False
    etiquetas.showPercent = False
    etiquetas.showBubbleSize = False
    etiquetas.dLblPos = "outEnd"
    etiquetas.numFmt = "#,##0"
    ch.dataLabels = etiquetas
    ch.height, ch.width = alto, ancho

    # Un color por barra. Hay que CONSTRUIR las propiedades gráficas: asignarlas
    # sobre un DataPoint recién creado no las serializa (quedan fuera del XML y
    # Excel pinta todas las barras del mismo azul por defecto).
    # `graphicalProperties` es solo un alias de lectura: el constructor exige `spPr`.
    serie = ch.series[0]
    serie.data_points = [
        DataPoint(idx=i, spPr=GraphicalProperties(
            solidFill=color, ln=LineProperties(noFill=True)))
        for i, color in enumerate(colores)
    ]
    return ch


# --- Carga ------------------------------------------------------------------

def cargar():
    if not INFORME.is_file():
        sys.exit(f"No existe {INFORME}. Corre antes: python migrar_cedulas_google.py informar ...")
    with INFORME.open(encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def diagnostico():
    """Cifras del directorio de Google. Vacío si no está el volcado."""
    if not VOLCADO.is_file():
        return {}
    cuentas = json.loads(VOLCADO.read_text(encoding="utf-8"))
    vacios = sum(1 for c in cuentas for e in c["externalIds"]
                 if not (e.get("value") or "").strip())
    valores = [(e.get("value") or "").strip() for c in cuentas for e in c["externalIds"]
               if (e.get("value") or "").strip()]
    ident = sum(1 for c in cuentas for e in c["externalIds"]
                if (e.get("customType") or e.get("type")) == "identificacion"
                and (e.get("value") or "").strip())
    return {
        "cuentas": len(cuentas),
        "con_externalids": sum(1 for c in cuentas if c["externalIds"]),
        "valores_vacios": vacios,
        "valores_reales": len(valores),
        "nueve_digitos": sum(1 for v in valores if v.isdigit() and len(v) == 9),
        "tipo_identificacion": ident,
    }


# --- Hojas ------------------------------------------------------------------

def hoja_resumen(wb, filas, conteo, grupos, diag):
    ws = wb.active
    ws.title = "Resumen"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for col in "BCDEFGHIJK":
        ws.column_dimensions[col].width = 11

    total = len(filas)
    listas = grupos["Listas para escribir"]
    revisar = grupos["Requieren revisión"]
    sin_cuenta = grupos["Sin cuenta en Google"]

    titulo(ws, "B2", "Migración de la cédula a Google Workspace", 18)
    ws["B3"] = "Informe gerencial · estado previo a la escritura"
    ws["B3"].font = Font(size=10, italic=True, color=TINTA_2)

    parrafo(ws, 5,
            "Hoy la cédula de las personas no vive en Google: de las "
            f"{mil(diag.get('cuentas', 0))} cuentas del dominio, solo "
            f"{diag.get('tipo_identificacion', 0)} la tienen registrada. Este proceso "
            "la escribe en la cuenta principal de cada persona, para que en adelante "
            "una cuenta pueda encontrarse por la cédula y Google sea la fuente de "
            "verdad.")
    parrafo(ws, 6,
            f"Se cruzaron {mil(total)} personas de la tabla institucional contra el "
            "directorio. El proceso NO ha escrito nada todavía: las cifras de abajo "
            "son el resultado del análisis, y la escritura solo alcanza al grupo "
            "verde.")

    kpi(ws, 8, 2, total, "Personas analizadas", TINTA)
    kpi(ws, 8, 4, listas, "Listas para escribir", BUENO)
    kpi(ws, 8, 6, sin_cuenta, "Sin cuenta en Google", AZUL)
    kpi(ws, 8, 8, revisar, "Requieren revisión humana", AVISO)
    kpi(ws, 8, 10, grupos["Ya vinculadas"], "Ya vinculadas", TINTA_2)

    # Datos del gráfico (a la derecha, fuera de las celdas combinadas y ocultos).
    ws["U1"] = "Personas"
    for i, g in enumerate(ORDEN_GRUPOS, start=2):
        ws.cell(row=i, column=20, value=g)
        ws.cell(row=i, column=21, value=grupos[g])
    datos = Reference(ws, min_col=21, min_row=1, max_row=1 + len(ORDEN_GRUPOS))
    cats = Reference(ws, min_col=20, min_row=2, max_row=1 + len(ORDEN_GRUPOS))
    ch = barra(datos, cats, "Qué va a pasar con cada persona",
               [COLOR_GRUPO[g] for g in ORDEN_GRUPOS], alto=6.5, ancho=17)
    ws.add_chart(ch, "B12")

    # Cobertura antes / después: la misma medida en dos momentos -> una sola serie.
    antes = diag.get("tipo_identificacion", 0)
    ws["U8"] = "Cuentas con cédula"
    ws["T9"], ws["U9"] = "Hoy", antes
    ws["T10"], ws["U10"] = "Tras la migración", antes + listas
    datos2 = Reference(ws, min_col=21, min_row=8, max_row=10)
    cats2 = Reference(ws, min_col=20, min_row=9, max_row=10)
    ch2 = barra(datos2, cats2, "Cuentas con la cédula registrada",
                [AZUL_CLARO, AZUL], alto=6.5, ancho=17)
    ws.add_chart(ch2, "B27")

    parrafo(ws, 26,
            f"El salto es de {antes} a {mil(antes + listas)} cuentas identificables por "
            "cédula. No llega al total del dominio porque hay cuentas que no "
            "corresponden a ninguna persona de la tabla (egresados antiguos, buzones "
            "de sistema) y porque las filas en revisión no se escriben.")

    for col in "TU":
        ws.column_dimensions[col].hidden = True
    return ws


def hoja_estados(wb, conteo):
    ws = wb.create_sheet("Estados")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    titulo(ws, "B2", "Detalle por estado")
    parrafo(ws, 3, "Cada persona recibe exactamente un estado. Los dos primeros los "
                   "escribe el proceso; el resto exige una decisión humana o no "
                   "aplica.")

    encabezado_tabla(ws, 5, ["Estado", "Personas", "Grupo", "Qué significa", "Qué hay que hacer"],
                     anchos=[22, 10, 20, 52, 44])
    fila = 6
    presentes = [e for e in ESTADOS if conteo.get(e)]
    presentes.sort(key=lambda e: -conteo[e])
    for e in presentes:
        grupo, color, significa, accion = ESTADOS[e]
        ws.cell(row=fila, column=2, value=e).font = Font(size=10, bold=True, color=color)
        c = ws.cell(row=fila, column=3, value=conteo[e])
        c.number_format = "#,##0"
        ws.cell(row=fila, column=4, value=grupo).font = Font(size=10, color=TINTA_2)
        for col, txt in ((5, significa), (6, accion)):
            cc = ws.cell(row=fila, column=col, value=txt)
            cc.font = Font(size=9, color=TINTA_2)
            cc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[fila].height = 30
        fila += 1

    # Gráfico solo de los estados que exigen revisión: son los accionables.
    # Datos auxiliares del gráfico, lejos de las celdas combinadas de los párrafos.
    rev = [(e, conteo[e]) for e in presentes if e in REVISION]
    if rev:
        ws["U1"] = "Personas"
        for i, (e, n) in enumerate(rev, start=2):
            ws.cell(row=i, column=20, value=e)
            ws.cell(row=i, column=21, value=n)
        datos = Reference(ws, min_col=21, min_row=1, max_row=1 + len(rev))
        cats = Reference(ws, min_col=20, min_row=2, max_row=1 + len(rev))
        colores = [ESTADOS[e][1] for e, _ in rev]
        # Ocho categorías de nombre largo: necesita más alto que los demás.
        ch = barra(datos, cats, "Casos que requieren revisión humana", colores,
                   alto=10, ancho=18)
        ws.add_chart(ch, f"B{fila + 2}")
        for col in "TU":
            ws.column_dimensions[col].hidden = True


def hoja_calidad(wb, diag):
    ws = wb.create_sheet("Calidad del dato")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    if not diag:
        ws["B2"] = "Falta pruebas/volcado_google.json"
        return

    titulo(ws, "B2", "Calidad del dato hoy en Google")
    parrafo(ws, 3, "Lo que se encontró al inspeccionar las "
                   f"{mil(diag['cuentas'])} cuentas del dominio. Explica por qué la "
                   "migración no podía apoyarse en lo que ya había.")

    hallazgos = [
        ("Cuentas en el dominio", diag["cuentas"],
         "Universo total, incluidas archivadas y buzones de sistema."),
        ("Cuentas con algún identificador externo", diag["con_externalids"],
         "Apenas el 2,5 % del dominio."),
        ("…de esos, con el valor VACÍO", diag["valores_vacios"],
         "Un campo creado pero nunca rellenado: no sirve para buscar."),
        ("Identificadores con valor real", diag["valores_reales"],
         "El 1,2 % del dominio. Era imposible migrar buscando por ellos."),
        ("Cédulas guardadas sin el cero inicial", diag["nueve_digitos"],
         "Se guardaron como número: '0925…' quedó como '925…'. Buscar la cédula "
         "real no las encuentra."),
        ("Cuentas con la cédula bien etiquetada", diag["tipo_identificacion"],
         "El resto usa una etiqueta distinta ('organization'), pensada para otra cosa."),
    ]

    encabezado_tabla(ws, 5, ["Hallazgo", "Cuentas", "Por qué importa"],
                     anchos=[46, 12, 76])
    fila = 6
    for texto, valor, porque in hallazgos:
        ws.cell(row=fila, column=2, value=texto).font = Font(size=10, color=TINTA)
        c = ws.cell(row=fila, column=3, value=valor)
        c.number_format = "#,##0"
        c.font = Font(size=10, bold=True, color=TINTA)
        cc = ws.cell(row=fila, column=4, value=porque)
        cc.font = Font(size=9, color=TINTA_2)
        cc.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[fila].height = 28
        fila += 1

    parrafo(ws, fila + 1,
            "Conclusión: el identificador externo existía en el papel, pero no en los "
            "datos. Por eso el emparejamiento se apoya en el correo institucional, "
            "que sí está en la tabla de personas y es exacto.")


def hoja_revision(wb, filas):
    ws = wb.create_sheet("Revisión")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2

    pendientes = [f for f in filas if f["estado"] in REVISION]
    titulo(ws, "B2", f"Casos a revisar ({len(pendientes)})")
    parrafo(ws, 3, "Ninguna de estas filas se escribe en Google. Requieren una "
                   "decisión: corregir la tabla de personas, o confirmar a mano qué "
                   "cuenta corresponde.")

    encabezado_tabla(ws, 5, ["Estado", "persona_id", "Cédula", "Nombre", "Correo (tabla)",
                             "Cuenta hallada", "Unidad", "Detalle"],
                     anchos=[22, 11, 13, 30, 32, 32, 30, 70])
    # Los críticos primero: son los que pueden corromper un dato.
    orden = {CRITICO: 0, AVISO: 1}
    pendientes.sort(key=lambda f: (orden.get(ESTADOS[f["estado"]][1], 2), f["estado"]))

    for i, f in enumerate(pendientes, start=6):
        color = ESTADOS[f["estado"]][1]
        ws.cell(row=i, column=2, value=f["estado"]).font = Font(size=9, bold=True, color=color)
        ws.cell(row=i, column=3, value=f["persona_id"])
        ws.cell(row=i, column=4, value=f["identificacion"])
        ws.cell(row=i, column=5, value=f"{f['nombres']} {f['apellidos']}".strip())
        ws.cell(row=i, column=6, value=f["emailinst"])
        ws.cell(row=i, column=7, value=f["email_google"])
        ws.cell(row=i, column=8, value=f["ou"])
        ws.cell(row=i, column=9, value=f["detalle"])
        for col in range(3, 10):
            ws.cell(row=i, column=col).font = Font(size=9, color=TINTA_2)
    ws.freeze_panes = "B6"
    ws.auto_filter.ref = f"B5:I{5 + len(pendientes)}"


def hoja_metodologia(wb, total, grupos):
    ws = wb.create_sheet("Metodología")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 110

    titulo(ws, "B2", "Cómo se emparejó cada persona")
    pasos = [
        ("1. Por cédula", "Si la cédula ya está escrita en alguna cuenta, esa cuenta "
                          "identifica a la persona. Se prueban las variantes con y sin "
                          "cero inicial."),
        ("2. Por correo institucional", "La llave exacta que ya está en la tabla de "
                                        "personas. Es de donde salen 14 de cada 15 "
                                        "emparejamientos."),
        ("3. Por nombre completo", "Solo como último recurso, y solo si el nombre es "
                                   "idéntico tras quitar acentos y ordenar las palabras. "
                                   "Si únicamente se parece, la fila va a revisión."),
    ]
    fila = 4
    for t, d in pasos:
        ws.cell(row=fila, column=2, value=t).font = Font(size=11, bold=True, color=TINTA)
        fila += 1
        c = ws.cell(row=fila, column=2, value=d)
        c.font = Font(size=10, color=TINTA_2)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[fila].height = 30
        fila += 2

    titulo(ws, f"B{fila}", "Cuál es la cuenta principal", 12)
    fila += 1
    c = ws.cell(row=fila, column=2,
                value="Una misma persona puede tener dos cuentas activas: por ejemplo, "
                      "docente y estudiante. La cédula se escribe en una sola, elegida "
                      "por este orden: Administrativos, luego Docentes, luego "
                      "Estudiantes. Si empatan, la de último acceso más reciente. "
                      "Nunca se escribe en cuentas archivadas, suspendidas, de sistema "
                      "ni en buzones compartidos.")
    c.font = Font(size=10, color=TINTA_2)
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[fila].height = 60
    fila += 2

    titulo(ws, f"B{fila}", "Garantías del proceso", 12)
    fila += 1
    for g in ("No se sobrescribe una cédula distinta: se reporta el conflicto.",
              "No se escribe si dos personas apuntan a la misma cuenta.",
              "No se escriben cédulas de relleno como 0000000000.",
              "Los identificadores externos que ya existían se conservan: la escritura "
              "lee, añade y vuelve a guardar la lista completa.",
              "La escritura solo ocurre con la opción --real; por defecto simula."):
        c = ws.cell(row=fila, column=2, value="•  " + g)
        c.font = Font(size=10, color=TINTA_2)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[fila].height = 26
        fila += 1


# --- Principal --------------------------------------------------------------

def main() -> None:
    filas = cargar()
    conteo = Counter(f["estado"] for f in filas)

    desconocidos = set(conteo) - set(ESTADOS)
    if desconocidos:
        sys.exit(f"Estados no catalogados en este informe: {sorted(desconocidos)}. "
                 "Añádelos a ESTADOS antes de generar el Excel.")

    grupos = defaultdict(int)
    for estado, n in conteo.items():
        grupos[ESTADOS[estado][0]] += n
    for g in ORDEN_GRUPOS:
        grupos.setdefault(g, 0)

    diag = diagnostico()

    wb = Workbook()
    hoja_resumen(wb, filas, conteo, grupos, diag)
    hoja_estados(wb, conteo)
    hoja_calidad(wb, diag)
    hoja_revision(wb, filas)
    hoja_metodologia(wb, len(filas), grupos)

    PRUEBAS.mkdir(exist_ok=True)
    salida = Path(sys.argv[1]) if len(sys.argv) > 1 else SALIDA
    try:
        wb.save(salida)
    except PermissionError:
        sys.exit(f"No se pudo escribir {salida}: ciérralo en Excel y vuelve a intentarlo.\n"
                 f"(o pasa otra ruta: python {Path(__file__).name} otra_ruta.xlsx)")

    print(f"{mil(len(filas))} personas -> {salida}")
    for g in ORDEN_GRUPOS:
        print(f"  {g:24} {mil(grupos[g]):>8}")


if __name__ == "__main__":
    main()
